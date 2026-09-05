import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\xenov\Downloads\cekbarkot\Laporan Grade Induk"
files = [
    ("2026-08-21", "21-8", "JUMAT", "grade tgl 21.xlsx"),
    ("2026-08-22", "22-8", "SABTU", "grade tgl 22.xlsx"),
    ("2026-08-23", "23-8", "MINGGU", "grade tgl 23.xlsx"),
    ("2026-08-24", "24-8", "SENIN", "grade tgl 24.xlsx"),
    ("2026-08-26", "26-8", "RABU", "grade induk tgl 26.xlsx"),
    ("2026-08-27", "27-8", "KAMIS", "buku_grade_induk_27-8-2026.xlsx"),
    ("2026-08-28", "28-8", "JUMAT", "grade tgl 28.xlsx"),
    ("2026-08-29", "29-8", "SABTU", "grade lengkap tgl 29.xlsx"),
    ("2026-08-30", "30-8", "MINGGU", "grade tgl 30.xlsx"),
    ("2026-08-31", "31-8", "SENIN", "grade tgl 31.xlsx"),
    ("2026-09-01", "1-9", "SELASA", "grade_tgl_1_september_sorted.xlsx"),
    ("2026-09-02", "2-9", "RABU", "Grade 2 sep.xlsx"),
    ("2026-09-03", "3-9", "KAMIS", "grade 3 sep.xlsx"),
    ("2026-09-04", "4-9", "JUMAT", "grade 4 sep.xlsx"),
]

all_items = []

for dt_str, tgl_short, hari, fname in files:
    fpath = os.path.join(folder, fname)
    wb_src = openpyxl.load_workbook(fpath, data_only=True)
    sheet_src = wb_src.active
    
    for r in range(5, sheet_src.max_row + 1):
        no_gud = sheet_src.cell(row=r, column=1).value
        grade1 = sheet_src.cell(row=r, column=2).value
        grade2 = sheet_src.cell(row=r, column=3).value
        barkot = sheet_src.cell(row=r, column=4).value
        kg = sheet_src.cell(row=r, column=5).value
        ket = sheet_src.cell(row=r, column=6).value
        
        if any(v is not None for v in [no_gud, grade1, grade2, barkot, kg, ket]):
            b_str = str(barkot).strip() if barkot is not None else ""
            if b_str and b_str.isdigit():
                try:
                    ng_int = int(no_gud)
                except:
                    ng_int = no_gud
                all_items.append({
                    "date": dt_str,
                    "tgl_short": tgl_short,
                    "hari": hari,
                    "file": fname,
                    "no_gud": ng_int,
                    "grade1": grade1,
                    "grade2": grade2,
                    "barkot": b_str,
                    "kg": kg,
                    "ket": ket if ket is not None else ""
                })

# Sort ascending by no_gud, then date
all_items.sort(key=lambda x: (x["no_gud"] if isinstance(x["no_gud"], int) else 999999, x["date"]))

print(f"Total rows to write: {len(all_items)}")

# Create Workbook
wb = openpyxl.Workbook()

# Style definitions
font_title = Font(name="Arial", size=20, bold=True, color="000000")
font_subtitle = Font(name="Arial", size=10, bold=True, color="000000")
font_header = Font(name="Arial", size=13, bold=True, color="FFFFFF")
font_data = Font(name="Arial", size=14, bold=False, color="000000")
font_total = Font(name="Arial", size=13, bold=True, color="FFFFFF")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_total = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

thin_side = Side(style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
bottom_thin_border = Border(bottom=thin_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# ==========================================
# SHEET 1: BUKU GRADE INDUK (Utama / Master)
# ==========================================
ws1 = wb.active
ws1.title = "Sheet1"

# Set column widths
ws1.column_dimensions['A'].width = 13.0
ws1.column_dimensions['B'].width = 9.0
ws1.column_dimensions['C'].width = 10.0
ws1.column_dimensions['D'].width = 21.0
ws1.column_dimensions['E'].width = 9.0
ws1.column_dimensions['F'].width = 28.57

# Row Heights
ws1.row_dimensions[1].height = 31.5
ws1.row_dimensions[2].height = 21.75
ws1.row_dimensions[3].height = 21.0
ws1.row_dimensions[4].height = 21.0

# Row 1: Title
ws1.merge_cells("A1:F1")
cell_a1 = ws1["A1"]
cell_a1.value = "BUKU GRADE INDUK"
cell_a1.font = font_title
cell_a1.alignment = align_center

# Row 2: Subtitle
ws1.merge_cells("A2:F2")
cell_a2 = ws1["A2"]
cell_a2.value = "TANGGAL: 21-8 s/d 4-9          HARI: ......................          TAHUN: 2026"
cell_a2.font = font_subtitle
cell_a2.alignment = align_center
for col in range(1, 7):
    ws1.cell(row=2, column=col).border = bottom_thin_border

# Row 3 & 4: Table Headers
header_merges = [
    ("A3:A4", "NO GUD"),
    ("B3:C4", "GRADE"),
    ("D3:D4", "BARKOT"),
    ("E3:E4", "KG"),
    ("F3:F4", "KET"),
]

for rng, val in header_merges:
    ws1.merge_cells(rng)
    top_left = ws1[rng.split(":")[0]]
    top_left.value = val

for r in [3, 4]:
    for c in range(1, 7):
        cell = ws1.cell(row=r, column=c)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

# Data Rows: Row 5 to len(all_items) + 4
start_row = 5
for idx, item in enumerate(all_items):
    curr_row = start_row + idx
    ws1.row_dimensions[curr_row].height = 20.25
    
    c_a = ws1.cell(row=curr_row, column=1, value=item["no_gud"])
    c_b = ws1.cell(row=curr_row, column=2, value=item["grade1"])
    c_c = ws1.cell(row=curr_row, column=3, value=item["grade2"])
    
    # Barkot formatted cleanly as text/string
    c_d = ws1.cell(row=curr_row, column=4, value=str(item["barkot"]))
    c_d.number_format = '@'
    
    c_e = ws1.cell(row=curr_row, column=5, value=item["kg"])
    c_f = ws1.cell(row=curr_row, column=6, value=item["ket"] if item["ket"] else None)
    
    for cell in [c_a, c_b, c_c, c_d, c_e, c_f]:
        cell.font = font_data
        cell.alignment = align_center
        cell.border = thin_border

# Total Row
last_data_row = start_row + len(all_items) - 1
total_row = last_data_row + 1
ws1.row_dimensions[total_row].height = 22.0

ws1.merge_cells(f"A{total_row}:D{total_row}")
cell_tot_lbl = ws1[f"A{total_row}"]
cell_tot_lbl.value = f"TOTAL ({len(all_items)} BAL)"

cell_tot_kg = ws1[f"E{total_row}"]
cell_tot_kg.value = f"=SUM(E{start_row}:E{last_data_row})"

cell_tot_ket = ws1[f"F{total_row}"]
cell_tot_ket.value = "KG"

for c in range(1, 7):
    cell = ws1.cell(row=total_row, column=c)
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = align_center
    cell.border = thin_border

# Enable gridlines
ws1.views.sheetView[0].showGridLines = True

# ==========================================
# SHEET 2: DATA DETAIL (+ TANGGAL)
# ==========================================
ws2 = wb.create_sheet(title="Data Lengkap + Tanggal")
ws2.views.sheetView[0].showGridLines = True

ws2.column_dimensions['A'].width = 8.0
ws2.column_dimensions['B'].width = 14.0
ws2.column_dimensions['C'].width = 12.0
ws2.column_dimensions['D'].width = 13.0
ws2.column_dimensions['E'].width = 10.0
ws2.column_dimensions['F'].width = 10.0
ws2.column_dimensions['G'].width = 20.0
ws2.column_dimensions['H'].width = 12.0
ws2.column_dimensions['I'].width = 25.0

ws2.merge_cells("A1:I1")
ws2["A1"].value = "REKAP DATA BAL DENGAN BARKOT (21 AGUSTUS - 4 SEPTEMBER 2026)"
ws2["A1"].font = Font(name="Arial", size=16, bold=True)
ws2["A1"].alignment = align_center

ws2.merge_cells("A2:I2")
ws2["A2"].value = f"Diurutkan berdasarkan No Gudang terkecil sampai terbesar | Total: {len(all_items)} Bal Berbarkot"
ws2["A2"].font = Font(name="Arial", size=10, italic=True)
ws2["A2"].alignment = align_center

headers2 = ["No", "Tanggal", "Hari", "No Gud", "Grade 1", "Grade 2", "Barkot", "Kg", "Sumber File"]
for c_idx, h_text in enumerate(headers2, 1):
    c = ws2.cell(row=4, column=c_idx, value=h_text)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = thin_border
ws2.row_dimensions[4].height = 24.0

for idx, item in enumerate(all_items, 1):
    r_idx = 4 + idx
    ws2.row_dimensions[r_idx].height = 20.0
    row_data = [
        idx,
        item["date"],
        item["hari"],
        item["no_gud"],
        item["grade1"],
        item["grade2"],
        str(item["barkot"]),
        item["kg"],
        item["file"]
    ]
    for c_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = Font(name="Arial", size=11)
        c.alignment = align_center if c_idx != 9 else align_left
        c.border = thin_border
        if c_idx == 7:
            c.number_format = '@'

# Total Row for ws2
tot2_row = 5 + len(all_items)
ws2.merge_cells(f"A{tot2_row}:G{tot2_row}")
ws2[f"A{tot2_row}"].value = f"TOTAL ({len(all_items)} BAL)"
ws2[f"H{tot2_row}"].value = f"=SUM(H5:H{tot2_row-1})"
ws2[f"I{tot2_row}"].value = "KG"

for c in range(1, 10):
    cell = ws2.cell(row=tot2_row, column=c)
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = align_center
    cell.border = thin_border

# ==========================================
# SHEET 3: RINGKASAN PER TANGGAL
# ==========================================
ws3 = wb.create_sheet(title="Ringkasan Per Tanggal")
ws3.views.sheetView[0].showGridLines = True

ws3.column_dimensions['A'].width = 8.0
ws3.column_dimensions['B'].width = 16.0
ws3.column_dimensions['C'].width = 14.0
ws3.column_dimensions['D'].width = 16.0
ws3.column_dimensions['E'].width = 16.0
ws3.column_dimensions['F'].width = 16.0
ws3.column_dimensions['G'].width = 20.0

ws3.merge_cells("A1:G1")
ws3["A1"].value = "RINGKASAN REKAP BAL BERBARKOT PER TANGGAL"
ws3["A1"].font = Font(name="Arial", size=16, bold=True)
ws3["A1"].alignment = align_center

ws3.merge_cells("A2:G2")
ws3["A2"].value = "Periode: 21 Agustus 2026 s/d 4 September 2026"
ws3["A2"].font = Font(name="Arial", size=10, italic=True)
ws3["A2"].alignment = align_center

headers3 = ["No", "Tanggal", "Hari", "Jumlah Bal", "Total Berat (Kg)", "Rata-rata Kg/Bal", "Rentang No Gudang"]
for c_idx, h_text in enumerate(headers3, 1):
    c = ws3.cell(row=4, column=c_idx, value=h_text)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = thin_border
ws3.row_dimensions[4].height = 24.0

from collections import defaultdict
summary = defaultdict(lambda: {"count": 0, "kg": 0, "no_guds": [], "hari": ""})
for it in all_items:
    d = it["date"]
    summary[d]["count"] += 1
    summary[d]["kg"] += it["kg"]
    summary[d]["no_guds"].append(it["no_gud"])
    summary[d]["hari"] = it["hari"]

for idx, dt in enumerate(sorted(summary.keys()), 1):
    r_idx = 4 + idx
    ws3.row_dimensions[r_idx].height = 20.0
    info = summary[dt]
    min_ng = min(info["no_guds"])
    max_ng = max(info["no_guds"])
    avg_kg = round(info["kg"] / info["count"], 2) if info["count"] > 0 else 0
    row_data = [
        idx,
        dt,
        info["hari"],
        info["count"],
        info["kg"],
        avg_kg,
        f"{min_ng} - {max_ng}"
    ]
    for c_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=r_idx, column=c_idx, value=val)
        c.font = Font(name="Arial", size=11)
        c.alignment = align_center
        c.border = thin_border

tot3_row = 5 + len(summary)
ws3.merge_cells(f"A{tot3_row}:C{tot3_row}")
ws3[f"A{tot3_row}"].value = "TOTAL KESELURUHAN"
ws3[f"D{tot3_row}"].value = f"=SUM(D5:D{tot3_row-1})"
ws3[f"E{tot3_row}"].value = f"=SUM(E5:E{tot3_row-1})"
ws3[f"F{tot3_row}"].value = f"=AVERAGE(F5:F{tot3_row-1})"
ws3[f"G{tot3_row}"].value = f"{min(it['no_gud'] for it in all_items)} - {max(it['no_gud'] for it in all_items)}"

for c in range(1, 8):
    cell = ws3.cell(row=tot3_row, column=c)
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = align_center
    cell.border = thin_border

# Save files
out_path_folder1 = os.path.join(folder, "buku_grade_induk_rekap_barkot_tgl_21_sd_31.xlsx")
out_path_folder2 = os.path.join(folder, "laporan_grade_induk_semua_barkot_tgl_21_sd_31.xlsx")
out_path_root = r"c:\Users\xenov\Downloads\cekbarkot\buku_grade_induk_rekap_barkot_tgl_21_sd_31.xlsx"

wb.save(out_path_folder1)
wb.save(out_path_folder2)
wb.save(out_path_root)

print(f"Saved successfully to:")
print(f"  1. {out_path_folder1}")
print(f"  2. {out_path_folder2}")
print(f"  3. {out_path_root}")
